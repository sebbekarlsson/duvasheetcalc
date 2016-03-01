from openpyxl import load_workbook
from duvasheetcalc.sheetcalculator import SheetCalculator
from duvasheetcalc.calcsaver import CalcSaver

class SheetLoader(object):

    def __init__(self):
        self.calculator = SheetCalculator()
        self.saver = CalcSaver()
        self.height = 31
        self.letters = ['A', 'B', 'C', 'D']


    def clean_days(self, days):
        new_days = []
        for day in days:
            if day['interval'] is not None:
                new_days.append(day)

        return new_days


    def load(self, filename, output):
        days = []

        wb = load_workbook(filename)
        ws = wb.active

        i = 2
        for row in range(0, self.height):
            day = {}
            for char in self.letters:
                coords = '{}{}'.format(char, i)
                cell = ws[coords]

                if char == 'A':
                    day['date'] = cell.value
                elif char == 'B':
                    day['interval'] = cell.value
                elif char == 'C':
                    day['hours'] = cell.value
                elif char == 'D':
                    day['ref'] = cell.value

            days.append(day)
            i+=1

        days = self.clean_days(days)
        calculated_days = self.calculator.calculate(days)

        return self.saver.save(days, output)
