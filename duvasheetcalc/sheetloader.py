from openpyxl import load_workbook
from duvasheetcalc.sheetcalculator import SheetCalculator
from duvasheetcalc.calcsaver import CalcSaver
from duvasheetcalc.utils import clean_days


class SheetLoader(object):

    def __init__(self):
        self.calculator = SheetCalculator()
        self.saver = CalcSaver()
        self.height = 31
        self.letters = ['A', 'B', 'C', 'D']


    def load(self, filename, output):
        days = []

        wb = load_workbook(filename)
        ws = wb.active

        i = 2
        for row in range(0, self.height):
            day = {}
            for char in self.letters:
                coords = '{}{}'.format(char, i)
                cell = ws[coords]

                if char == 'A':
                    if cell.value is not None:
                        day['date'] = cell.value.strftime("%d-%B-%Y")
                elif char == 'B':
                    day['interval'] = cell.value
                elif char == 'C':
                    day['hours'] = cell.value
                elif char == 'D':
                    day['ref'] = cell.value

            days.append(day)
            i+=1

        days = clean_days(days)
        calculations = self.calculator.calculate(days)

        return self.saver.save(calculations, output)
